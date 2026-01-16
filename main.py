# main.py
# Excel-only ODC generator (NO merged cells). Grid-uniform layout (columns B:Z width=2.5).
# FastAPI endpoint: POST /generate-odc-excel
#
# Python 3.11 + openpyxl
#
# Notes:
# - We simulate "merged" headers using Alignment(horizontal="centerContinuous") a.k.a. Center Across Selection.
# - Layout is drawn by painting rectangles (fills/borders) and writing text across ranges.

from __future__ import annotations

from io import BytesIO
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI
from fastapi.responses import Response
from pydantic import BaseModel, Field

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# -----------------------------------------------------------------------------
# FastAPI
# -----------------------------------------------------------------------------
app = FastAPI(title="ODCs - Excel (No Merges)")


# -----------------------------------------------------------------------------
# Models
# -----------------------------------------------------------------------------
class ODCItem(BaseModel):
    concept: str
    unit_cost: float
    units: float

    @property
    def subtotal(self) -> float:
        return float(self.unit_cost) * float(self.units)


class ODCRequest(BaseModel):
    odc_number: str = Field(..., examples=["RI-02497"])
    issue_date: Optional[str] = Field(None, examples=["17 nov 2025"])  # keep as string for exact formatting
    supplier: str
    service: str
    project: str

    bill_to_name: str = Field(..., examples=["ASESORES GLOBALES CORPORATIVOS"])
    bill_to_rfc: str = Field(..., examples=["AGC051117MX5"])
    bill_to_address_1: str = Field(..., examples=["Peregrinos 24, Colinas del Sur,"])
    bill_to_address_2: str = Field(..., examples=["Álvaro Obregón, CP. 01430, CDMX"])

    items: List[ODCItem] = Field(default_factory=list)


# -----------------------------------------------------------------------------
# Grid + Ranges (based on your screenshot)
# -----------------------------------------------------------------------------
GRID_COL_START = "B"
GRID_COL_END = "Z"
GRID_ROW_TOP = 2

BASE_COL_WIDTH = 2.5
BASE_ROW_HEIGHT = 18  # uniform baseline; specific rows override below

# Convenience: convert "B".."Z" to indices
C0 = column_index_from_string(GRID_COL_START)
C1 = column_index_from_string(GRID_COL_END)

# Major blocks (row, col, row, col) in grid coordinates
R_BANNER = (2, C0, 4, C1)          # B2:Z4
R_ODC_BOX = (3, column_index_from_string("T"), 3, C1)  # T3:Z3

R_LEFT_BLOCK = (5, C0, 9, column_index_from_string("O"))      # B5:O9
R_RIGHT_BLOCK = (5, column_index_from_string("Q"), 9, C1)     # Q5:Z9

R_ITEMS_HEADER = (11, C0, 11, C1)   # B11:Z11
# items body: starts at row 12; ends dynamically

# Column splits for items table (approx from screenshot)
# Concept: B:N | Unit cost: O:S | Units: T:V | Subtotal: W:Z
COL_CONCEPT = (column_index_from_string("B"), column_index_from_string("N"))
COL_UNIT_COST = (column_index_from_string("O"), column_index_from_string("S"))
COL_UNITS = (column_index_from_string("T"), column_index_from_string("V"))
COL_SUBTOTAL = (column_index_from_string("W"), column_index_from_string("Z"))

# Left info block label/value split
LEFT_LABEL = (column_index_from_string("B"), column_index_from_string("E"))  # B:E
LEFT_VALUE = (column_index_from_string("F"), column_index_from_string("O"))  # F:O
LEFT_DIVIDER_COL = column_index_from_string("E")  # vertical divider at end of labels


# -----------------------------------------------------------------------------
# Styles
# -----------------------------------------------------------------------------
# Colors (tuned to screenshot vibe)
COLOR_TEAL = "0F3B4A"        # dark teal
COLOR_TEAL_2 = "0D4A60"      # slightly brighter for table header
COLOR_LIGHT_GRAY = "EFEFEF"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"
COLOR_ACCENT_LABEL = "0D4A60"
COLOR_RED = "D50000"

FILL_TEAL = PatternFill("solid", fgColor=COLOR_TEAL)
FILL_TEAL_2 = PatternFill("solid", fgColor=COLOR_TEAL_2)
FILL_GRAY = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY)
FILL_WHITE = PatternFill("solid", fgColor=COLOR_WHITE)

SIDE_THIN = Side(style="thin", color="6B6B6B")
BORDER_THIN = Border(left=SIDE_THIN, right=SIDE_THIN, top=SIDE_THIN, bottom=SIDE_THIN)

SIDE_DIV = Side(style="thin", color="8A8A8A")
BORDER_RIGHT_DIV = Border(right=SIDE_DIV)
BORDER_LEFT_DIV = Border(left=SIDE_DIV)

FONT_SAPIENCE = Font(name="Calibri", size=34, bold=True, color="FFFFFF")
FONT_TAGLINE = Font(name="Calibri", size=12, color="CFE3EA")
FONT_LABEL = Font(name="Calibri", size=14, bold=True, color=COLOR_ACCENT_LABEL)
FONT_VALUE = Font(name="Calibri", size=16, color=COLOR_BLACK)

FONT_FACTURAR = Font(name="Calibri", size=22, bold=True, color=COLOR_ACCENT_LABEL)
FONT_BILL_TO_BOLD = Font(name="Calibri", size=16, bold=True, color=COLOR_BLACK)
FONT_BILL_TO = Font(name="Calibri", size=16, color=COLOR_BLACK)

FONT_TABLE_HEADER = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_TABLE_BODY = Font(name="Calibri", size=16, color=COLOR_BLACK)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Center across selection (no merges)
ALIGN_CENTER_ACROSS = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)


# -----------------------------------------------------------------------------
# Helpers (NO merges)
# -----------------------------------------------------------------------------
def set_uniform_grid(ws: Worksheet) -> None:
    # Columns B:Z all same width
    for col in range(C0, C1 + 1):
        ws.column_dimensions[_col_letter(col)].width = BASE_COL_WIDTH

    # Basic row heights (we set a reasonable span)
    for r in range(1, 200):
        ws.row_dimensions[r].height = BASE_ROW_HEIGHT

    # Custom heights to match screenshot proportions
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[11].height = 28  # table header slightly taller


def _col_letter(col_idx: int) -> str:
    # openpyxl get_column_letter is fine, but avoid extra import
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx)


def fill_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, fill: PatternFill) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill


def border_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, border: Border) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def outline_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, border: Border) -> None:
    # simple outline: apply border to all cells on perimeter
    for c in range(c1, c2 + 1):
        ws.cell(row=r1, column=c).border = border
        ws.cell(row=r2, column=c).border = border
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=c1).border = border
        ws.cell(row=r, column=c2).border = border


def vline_right(ws: Worksheet, col: int, r1: int, r2: int, side: Side = SIDE_DIV) -> None:
    for r in range(r1, r2 + 1):
        cell = ws.cell(row=r, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=side,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )


def write_across(
    ws: Worksheet,
    row: int,
    col_start: int,
    col_end: int,
    value: Any,
    *,
    font: Optional[Font] = None,
    fill: Optional[PatternFill] = None,
    align: Alignment = ALIGN_CENTER_ACROSS,
    number_format: Optional[str] = None,
) -> None:
    # Put the value only in the left cell
    cell0 = ws.cell(row=row, column=col_start)
    cell0.value = value

    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = align
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format


def write_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    font: Optional[Font] = None,
    align: Alignment = ALIGN_LEFT,
    fill: Optional[PatternFill] = None,
    number_format: Optional[str] = None,
) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = align
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


# -----------------------------------------------------------------------------
# Drawing sections (NO merges)
# -----------------------------------------------------------------------------
def draw_banner(ws: Worksheet, data: ODCRequest) -> None:
    r1, c1, r2, c2 = R_BANNER
    fill_range(ws, r1, c1, r2, c2, FILL_TEAL)

    # "SAPIENCE" and tagline (left)
    write_cell(ws, 3, column_index_from_string("B"), "SAPIENCE", font=FONT_SAPIENCE, align=ALIGN_LEFT, fill=FILL_TEAL)
    write_cell(ws, 4, column_index_from_string("B"), "Human Insights Strategy", font=FONT_TAGLINE, align=ALIGN_LEFT, fill=FILL_TEAL)

    # ODC box (top-right) - draw rectangle + divider
    br1, bc1, br2, bc2 = R_ODC_BOX
    fill_range(ws, br1, bc1, br2, bc2, FILL_GRAY)
    border_range(ws, br1, bc1, br2, bc2, BORDER_THIN)

    # divider roughly between label and value (use W as divider)
    divider_col = column_index_from_string("W")
    vline_right(ws, divider_col, br1, br2, SIDE_THIN)

    # label "ODC #:" on left part
    write_across(
        ws, br1,
        bc1, divider_col,
        "ODC #:",
        font=Font(name="Calibri", size=16, bold=True, color=COLOR_ACCENT_LABEL),
        fill=FILL_GRAY,
        align=ALIGN_CENTER_ACROSS
    )

    # value in red on right part
    write_across(
        ws, br1,
        divider_col + 1, bc2,
        data.odc_number,
        font=Font(name="Calibri", size=18, bold=True, color=COLOR_RED),
        fill=FILL_GRAY,
        align=ALIGN_CENTER_ACROSS
    )


def draw_left_info(ws: Worksheet, data: ODCRequest) -> None:
    r1, c1, r2, c2 = R_LEFT_BLOCK

    # Zebra rows like screenshot: 5,7,9 grey; 6,8 white
    for row in range(r1, r2 + 1):
        fill = FILL_GRAY if row in (5, 7, 9) else FILL_WHITE
        fill_range(ws, row, c1, row, c2, fill)

    # Divider line between label and value (end of E)
    vline_right(ws, LEFT_DIVIDER_COL, r1, r2, SIDE_DIV)

    # Labels
    labels = [
        ("ODC #", 5),
        ("FECHA:", 6),
        ("PROVEEDOR:", 7),
        ("SERVICIO:", 8),
        ("PROYECTO:", 9),
    ]
    for text, row in labels:
        # right-aligned label across B:E
        write_across(
            ws, row,
            LEFT_LABEL[0], LEFT_LABEL[1],
            text,
            font=FONT_LABEL,
            fill=ws.cell(row=row, column=LEFT_LABEL[0]).fill,  # keep zebra
            align=Alignment(horizontal="right", vertical="center", wrap_text=True),
        )

    # Values
    issue = data.issue_date or ""  # keep string formatting given by client
    values = [
        (data.odc_number, 5),
        (issue, 6),
        (data.supplier, 7),
        (data.service, 8),
        (data.project, 9),
    ]
    for val, row in values:
        # left-aligned across F:O
        write_across(
            ws, row,
            LEFT_VALUE[0], LEFT_VALUE[1],
            val,
            font=FONT_VALUE,
            fill=ws.cell(row=row, column=LEFT_VALUE[0]).fill,
            align=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )


def draw_bill_to(ws: Worksheet, data: ODCRequest) -> None:
    # Minimal styling; text positions match screenshot vibe
    r1, c1, r2, c2 = R_RIGHT_BLOCK
    # keep background white
    fill_range(ws, r1, c1, r2, c2, FILL_WHITE)

    # "FACTURAR A:" large teal
    write_across(
        ws, 5,
        c1, c2,
        "FACTURAR A:",
        font=FONT_FACTURAR,
        fill=FILL_WHITE,
        align=Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True),
    )

    # Lines below (left aligned inside the block)
    # place starting at Q6 (c1)
    write_cell(ws, 6, c1, data.bill_to_name, font=FONT_BILL_TO_BOLD, align=ALIGN_LEFT)
    write_cell(ws, 7, c1, f"RFC: {data.bill_to_rfc}", font=FONT_BILL_TO_BOLD, align=ALIGN_LEFT)
    write_cell(ws, 8, c1, data.bill_to_address_1, font=FONT_BILL_TO, align=ALIGN_LEFT)
    write_cell(ws, 9, c1, data.bill_to_address_2, font=FONT_BILL_TO, align=ALIGN_LEFT)


def draw_items_table(ws: Worksheet, items: List[ODCItem]) -> int:
    # Header
    hr, c1, _, c2 = R_ITEMS_HEADER
    fill_range(ws, hr, c1, hr, c2, FILL_TEAL_2)
    border_range(ws, hr, c1, hr, c2, BORDER_THIN)

    # Vertical dividers for header + body
    # We'll add borders in a full range after body is known; for now header titles:
    write_across(ws, hr, COL_CONCEPT[0], COL_CONCEPT[1], "Concepto", font=FONT_TABLE_HEADER, fill=FILL_TEAL_2)
    write_across(ws, hr, COL_UNIT_COST[0], COL_UNIT_COST[1], "Costo unitario", font=FONT_TABLE_HEADER, fill=FILL_TEAL_2)
    write_across(ws, hr, COL_UNITS[0], COL_UNITS[1], "Unidades", font=FONT_TABLE_HEADER, fill=FILL_TEAL_2)
    write_across(ws, hr, COL_SUBTOTAL[0], COL_SUBTOTAL[1], "Subtotal", font=FONT_TABLE_HEADER, fill=FILL_TEAL_2)

    # Body starts at row 12
    start_row = hr + 1
    row = start_row

    # If no items, still draw 1 empty row like a template
    if not items:
        items = [ODCItem(concept="", unit_cost=0.0, units=0.0)]

    for idx, it in enumerate(items):
        # zebra: row 13,15,... light gray (as screenshot)
        fill = FILL_GRAY if (row % 2 == 1) else FILL_WHITE
        fill_range(ws, row, c1, row, c2, fill)

        # Concept (left)
        write_across(
            ws, row, COL_CONCEPT[0], COL_CONCEPT[1],
            it.concept,
            font=FONT_TABLE_BODY,
            fill=fill,
            align=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        # Unit cost (center)
        write_across(
            ws, row, COL_UNIT_COST[0], COL_UNIT_COST[1],
            it.unit_cost if it.unit_cost else "",
            font=FONT_TABLE_BODY,
            fill=fill,
            align=ALIGN_CENTER_ACROSS,
            number_format='"$"#,##0',
        )
        # Units (center)
        write_across(
            ws, row, COL_UNITS[0], COL_UNITS[1],
            it.units if it.units else "",
            font=FONT_TABLE_BODY,
            fill=fill,
            align=ALIGN_CENTER_ACROSS,
            number_format='0',
        )
        # Subtotal (center)
        write_across(
            ws, row, COL_SUBTOTAL[0], COL_SUBTOTAL[1],
            it.subtotal if it.subtotal else "",
            font=FONT_TABLE_BODY,
            fill=fill,
            align=ALIGN_CENTER_ACROSS,
            number_format='"$"#,##0',
        )

        row += 1

    end_row = row - 1

    # Apply borders to the whole table (header + body)
    border_range(ws, hr, c1, end_row, c2, BORDER_THIN)

    return end_row


def set_print_settings(ws: Worksheet, last_row: int) -> None:
    # Paper & scaling: tuned for stable PDF export from Excel
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # Print area: from B2 to Z<last_row> (table end)
    ws.print_area = f"{GRID_COL_START}{GRID_ROW_TOP}:{GRID_COL_END}{last_row}"


# -----------------------------------------------------------------------------
# Workbook builder
# -----------------------------------------------------------------------------
def build_odc_workbook(data: ODCRequest) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "ODC"

    set_uniform_grid(ws)

    # Base background (white across an ample region)
    fill_range(ws, 1, 1, 200, 60, FILL_WHITE)

    draw_banner(ws, data)
    draw_left_info(ws, data)
    draw_bill_to(ws, data)
    last_row = draw_items_table(ws, data.items)

    set_print_settings(ws, last_row)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------------------------------------------------------
# Endpoint
# -----------------------------------------------------------------------------
@app.post("/generate-odc-excel")
def generate_odc_excel(payload: ODCRequest):
    wb = build_odc_workbook(payload)
    xlsx_bytes = workbook_to_bytes(wb)

    filename = f"ODC_{payload.odc_number}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
